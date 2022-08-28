# write a program that process thousands of spreadsheets in a second
# i.e. Automation with python
"""
Project : Process Excel Spreadsheets
Using openpyxl , we can do any type of automation in a excel sheet
Here we just correct some values and add it into a new column
we can do any type of excel operations here
Also , here add a BarChart
"""
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_excel_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 2
        cell2 = sheet.cell(row, 4)
        cell2.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)


if __name__ == "__main__":
    print("Process Excel Spreadsheets")
    filename = input("Enter a file name with .xlsx extention ")
    process_excel_workbook(filename)
